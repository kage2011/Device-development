#!/usr/bin/env python3
import argparse, datetime as dt, json, os, re, subprocess, tempfile
from pathlib import Path
from openpyxl import load_workbook
from PIL import Image


def run(cmd):
    return subprocess.check_output(cmd, stderr=subprocess.STDOUT, text=True)


def ocr_text_for_rotation(img_path: str, rot: int, td: str) -> tuple[int, str]:
    rimg = os.path.join(td, f"r{rot}.png")
    Image.open(img_path).rotate(rot, expand=True).save(rimg)
    out = os.path.join(td, f"ocr_{rot}")
    subprocess.run(["tesseract", rimg, out, "-l", "jpn", "--oem", "1", "--psm", "6"], check=False)
    txt_path = out + ".txt"
    txt = open(txt_path, encoding="utf-8", errors="ignore").read() if os.path.exists(txt_path) else ""
    score = len(re.findall(r"[ぁ-んァ-ヶ一-龥]", txt))
    return score, txt


def best_ocr(pdf_path: str):
    with tempfile.TemporaryDirectory() as td:
        base = os.path.join(td, "page")
        run(["pdftoppm", "-png", "-r", "300", pdf_path, base])
        img = f"{base}-1.png"
        best = (-1, 0, "")
        for rot in [0, 90, 180, 270]:
            score, txt = ocr_text_for_rotation(img, rot, td)
            if score > best[0]:
                best = (score, rot, txt)
        return best[1], best[2]


def apply_rules(text: str, rules_path: str) -> str:
    if not rules_path or not os.path.exists(rules_path):
        return text
    rules = json.load(open(rules_path, encoding="utf-8"))
    for k, v in rules.get("replacements", {}).items():
        text = text.replace(k, v)
    return text


def pick(patterns, text):
    for p in patterns:
        m = re.search(p, text, re.MULTILINE)
        if m:
            return m.group(1).strip()
    return None


def parse_fields(text: str, pdf_name: str):
    date = None
    m = re.search(r"(20\d{2})[^\d]?(\d{1,2})[^\d]?(\d{1,2})", pdf_name)
    if m:
        date = dt.datetime(int(m.group(1)), int(m.group(2)), int(m.group(3)))

    equip_no = pick([
        r"設備No\s*[:：]?\s*([A-Za-z0-9\-]+)",
        r"設備番号\s*[:：]?\s*([A-Za-z0-9\-]+)",
    ], text)
    m2 = re.search(r"(NCD|NCMC|NCP|NC)[\-\s]?(\d{3})", pdf_name, re.IGNORECASE)
    if m2:
        equip_no = f"{m2.group(1).upper()}-{m2.group(2)}"

    occurrence = pick([r"発生[^\n]*\s(段取り時|立ち上げ時|加工中|待機時|起動時)", r"発生時\s*[:：]\s*(.+)"] , text)
    alarm_where = pick([r"アラーム内容/どこが\s*[:：]\s*(.+)", r"どこが\s*[:：]\s*(.+)"] , text)
    symptom = pick([r"症状\s*[:：]\s*(.+)", r"症状[^\n]*\n(.+)", r"症状\s*\(.*\)\s*\n(.+)"] , text)
    cause = pick([r"原因\s*[:：]\s*(.+)", r"原因[^\n]*\n(.+)"] , text)
    action = pick([r"処置\s*[:：]\s*(.+)", r"処置[^\n]*\n(.+)"] , text)

    return {"date": date, "equip_no": equip_no, "occurrence": occurrence, "alarm_where": alarm_where,
            "symptom": symptom, "cause": cause, "action": action}


def generate_captures(pdf_path: str, rotation: int, out_dir: str):
    os.makedirs(out_dir, exist_ok=True)
    stem = Path(pdf_path).stem
    base_png = os.path.join(out_dir, f"{stem}_page1.png")
    run(["pdftoppm", "-png", "-singlefile", "-r", "300", pdf_path, base_png[:-4]])

    img = Image.open(base_png)
    rot_img = img.rotate(rotation, expand=True)
    rot_path = os.path.join(out_dir, f"{stem}_page1_rot{rotation}.png")
    rot_img.save(rot_path)

    outbase = os.path.join(out_dir, f"{stem}_ocr")
    subprocess.run(["tesseract", rot_path, outbase, "-l", "jpn", "--oem", "1", "--psm", "6", "tsv"], check=False)
    tsv_path = outbase + ".tsv"

    keywords = {
        "date": ["修理依頼日", "年月日"],
        "equip_no": ["設備No", "設備番号"],
        "occurrence": ["発生発見", "発生時"],
        "alarm_where": ["アラーム", "どこが"],
        "symptom": ["症状"],
        "cause": ["原因"],
        "action": ["処置"],
    }
    captures = {"page": rot_path}

    if not os.path.exists(tsv_path):
        return captures

    words = []
    with open(tsv_path, encoding="utf-8", errors="ignore") as f:
        lines = f.read().splitlines()
    for ln in lines[1:]:
        cols = ln.split("\t")
        if len(cols) < 12:
            continue
        txt = cols[11].strip()
        if not txt:
            continue
        try:
            x, y, w, h = map(int, cols[6:10])
        except Exception:
            continue
        words.append((txt, x, y, w, h))

    W, H = rot_img.size
    for field, kws in keywords.items():
        hit = None
        for txt, x, y, w, h in words:
            if any(k in txt for k in kws):
                hit = (x, y, w, h)
                break
        if not hit:
            continue
        x, y, w, h = hit
        # keyword周辺を広めに切り出し
        left = max(0, x - 80)
        top = max(0, y - 30)
        right = min(W, x + 900)
        bottom = min(H, y + 260)
        crop = rot_img.crop((left, top, right, bottom))
        cpath = os.path.join(out_dir, f"{stem}_{field}.png")
        crop.save(cpath)
        captures[field] = cpath

    return captures


def append_excel(xlsx_path: str, fields: dict, source_pdf: str):
    wb = load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    equip_prefix = None
    if fields["equip_no"] and "-" in fields["equip_no"]:
        equip_prefix = fields["equip_no"].split("-")[0]

    row = [
        fields["date"], None, '-', '-', (fields.get("alarm_where") or '修理報告書OCR取込'), None, fields.get("occurrence"),
        fields["symptom"], fields["cause"], fields["action"],
        None, None, equip_prefix, fields["equip_no"], None, None,
        f"元資料: {os.path.basename(source_pdf)} / OCR自動追記", None,
    ]
    ws.append(row)
    wb.save(xlsx_path)
    return ws.max_row


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdf", required=True)
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--rules", default="")
    ap.add_argument("--print-only", action="store_true")
    ap.add_argument("--review", action="store_true", help="Appendせず、確認質問用にキャプチャ出力")
    ap.add_argument("--capture-dir", default="/home/kage/share/ocr_captures")
    # 確認後に手動で確定値を上書きするためのオプション
    ap.add_argument("--symptom", default=None)
    ap.add_argument("--cause", default=None)
    ap.add_argument("--action", default=None)
    ap.add_argument("--equip-no", default=None)
    ap.add_argument("--occurrence", default=None)
    ap.add_argument("--alarm-where", default=None)
    args = ap.parse_args()

    rotation, txt = best_ocr(args.pdf)
    txt = apply_rules(txt, args.rules)
    fields = parse_fields(txt, os.path.basename(args.pdf))

    print("[OCR EXTRACT]")
    print(f"- rotation: {rotation}")
    for k, v in fields.items():
        print(f"- {k}: {v}")

    captures = generate_captures(args.pdf, rotation, args.capture_dir)
    print("[CAPTURES]")
    for k, v in captures.items():
        print(f"- {k}: {v}")

    # 手動確定値で上書き
    if args.symptom:
        fields["symptom"] = args.symptom
    if args.cause:
        fields["cause"] = args.cause
    if args.action:
        fields["action"] = args.action
    if args.equip_no:
        fields["equip_no"] = args.equip_no
    if args.occurrence:
        fields["occurrence"] = args.occurrence
    if args.alarm_where:
        fields["alarm_where"] = args.alarm_where

    if args.print_only or args.review:
        print("\n[確認依頼テンプレ]")
        print("以下の形式で、OK か 修正文字列を返してください。")
        checks = [
            ("日時", fields.get("date"), captures.get("date")),
            ("設備番号", fields.get("equip_no"), captures.get("equip_no")),
            ("アラーム内容/どこが", fields.get("alarm_where"), captures.get("alarm_where")),
            ("発生時", fields.get("occurrence"), captures.get("occurrence")),
            ("症状/どう", fields.get("symptom"), captures.get("symptom")),
            ("原因/部位", fields.get("cause"), captures.get("cause")),
            ("処置", fields.get("action"), captures.get("action")),
        ]
        for label, val, cap in checks:
            print(f"- {label}: {val} でいい？")
            if cap:
                print(f"  capture: {cap}")
        return

    row = append_excel(args.xlsx, fields, args.pdf)
    print(f"[OK] appended row={row} to {args.xlsx}")


if __name__ == "__main__":
    main()
